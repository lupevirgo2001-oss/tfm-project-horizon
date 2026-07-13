import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from matplotlib.ticker import FuncFormatter, PercentFormatter
from .config import OUTPUT_DIR, TABLES_DIR, FIGURES_DIR, PROCESSED_DIR
from .metrics import summarize_backtest

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DYNAMIC_STRATEGIES = ["LSTM", "GRU", "GARCH", "EWMA"]
MAIN_STRATEGIES = ["GARCH", "EWMA", "STATIC", "LSTM", "GRU"]
ALL_STRATEGY_ORDER = ["GARCH", "EWMA", "STATIC", "LSTM", "GRU"]


def compute_monthly_fx_returns(features_df: pd.DataFrame) -> pd.DataFrame:
    monthly_price = features_df["EURUSD"].resample("ME").last()
    ret = monthly_price.pct_change().dropna()

    return pd.DataFrame(
        {
            "month": ret.index.to_period("M").astype(str),
            "fx_return_month": ret.values,
        }
    )


def _prepare_backtest(backtest_df: pd.DataFrame) -> pd.DataFrame:
    tmp = backtest_df.copy()
    tmp["month_dt"] = pd.PeriodIndex(tmp["month"], freq="M").to_timestamp()
    tmp = tmp.sort_values(["strategy", "month_dt"]).reset_index(drop=True)
    tmp["cum_net_pnl"] = tmp.groupby("strategy")["net_pnl"].cumsum()
    tmp["abs_residual_exposure"] = tmp["residual_exposure"].abs()
    tmp["hedge_ratio_pct"] = (tmp["hedge_ratio"] * 100).round(0).astype(int)
    return tmp


def _sort_strategies(df: pd.DataFrame, col: str = "strategy") -> pd.DataFrame:
    order_map = {name: i for i, name in enumerate(ALL_STRATEGY_ORDER)}
    out = df.copy()
    out["_order"] = out[col].map(order_map).fillna(99)
    out = out.sort_values(["_order", col]).drop(columns="_order")
    return out


def _format_number_columns(df: pd.DataFrame, decimals: int = 2) -> pd.DataFrame:
    out = df.copy()
    numeric_cols = out.select_dtypes(include="number").columns
    out[numeric_cols] = out[numeric_cols].round(decimals)
    return out


def _strategy_comment(strategy: str) -> str:
    comments = {
        "GARCH": "Benchmark econométrico clásico de volatilidad.",
        "EWMA": "Benchmark RiskMetrics/industria.",
        "STATIC": "Cobertura fija de referencia corporativa.",
        "LSTM": "Modelo recurrente Deep Learning basado en arquitectura LSTM.",
        "GRU": "Modelo recurrente Deep Learning basado en arquitectura GRU.",
    }
    return comments.get(strategy, "")


def _strategy_type(strategy: str) -> str:
    if strategy in ["LSTM", "GRU"]:
        return "Deep Learning"
    if strategy in ["GARCH", "EWMA"]:
        return "Benchmark volatilidad"
    if strategy == "STATIC":
        return "Benchmark estático"
    return "Otro"


def _export_backtest_tables(backtest_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    tmp = _prepare_backtest(backtest_df)

    summary = summarize_backtest(backtest_df).copy()

    summary["rank_cum_net_pnl"] = summary["cum_net_pnl"].rank(
        ascending=False, method="dense"
    ).astype(int)

    summary["tipo_estrategia"] = summary["strategy"].apply(_strategy_type)
    summary["comentario_resultado"] = summary["strategy"].apply(_strategy_comment)

    summary = summary.sort_values("rank_cum_net_pnl").reset_index(drop=True)

    # Tabla ejecutiva de resultados
    executive = summary[
        [
            "rank_cum_net_pnl",
            "strategy",
            "tipo_estrategia",
            "cum_net_pnl",
            "mean_net_pnl",
            "std_net_pnl",
            "residual_exposure_std",
            "max_drawdown",
            "comentario_resultado",
        ]
    ].copy()

    executive = executive.rename(
        columns={
            "rank_cum_net_pnl": "Ranking",
            "strategy": "Estrategia",
            "tipo_estrategia": "Tipo",
            "cum_net_pnl": "PnL proxy neto acumulado",
            "mean_net_pnl": "PnL proxy neto medio mensual",
            "std_net_pnl": "Volatilidad PnL proxy mensual",
            "residual_exposure_std": "Desv. exposición residual",
            "max_drawdown": "Máx. drawdown",
            "comentario_resultado": "Comentario",
        }
    )

    executive = _format_number_columns(executive, decimals=2)

    summary.to_csv(TABLES_DIR / "backtest_summary.csv", index=False)
    executive.to_csv(TABLES_DIR / "tabla_1_resumen_backtest.csv", index=False)

    # Perfil operativo de cobertura
    operational = (
        tmp.groupby("strategy")
        .agg(
            Meses=("month", "nunique"),
            Hedge_ratio_promedio=("hedge_ratio", "mean"),
            Hedge_ratio_minimo=("hedge_ratio", "min"),
            Hedge_ratio_maximo=("hedge_ratio", "max"),
            Exposicion_residual_abs_promedio=("abs_residual_exposure", "mean"),
            Costo_transaccional_promedio=("transaction_cost", "mean"),
            Costo_transaccional_total=("transaction_cost", "sum"),
        )
        .reset_index()
        .rename(columns={"strategy": "Estrategia"})
    )

    dist_counts = (
        tmp.groupby(["strategy", "hedge_ratio_pct"])
        .size()
        .reset_index(name="Meses_en_tramo")
    )

    dist_total = tmp.groupby("strategy")["month"].nunique().reset_index(name="Total_meses")
    dist_counts = dist_counts.merge(dist_total, on="strategy", how="left")
    dist_counts["Porcentaje_meses"] = dist_counts["Meses_en_tramo"] / dist_counts["Total_meses"]

    dist_pivot = (
        dist_counts.pivot_table(
            index="strategy",
            columns="hedge_ratio_pct",
            values="Porcentaje_meses",
            fill_value=0.0,
        )
        .reset_index()
        .rename(columns={"strategy": "Estrategia"})
    )

    for col in [30, 60, 90, 100]:
        if col not in dist_pivot.columns:
            dist_pivot[col] = 0.0

    dist_pivot = dist_pivot.rename(
        columns={
            30: "% meses HR 30%",
            60: "% meses HR 60%",
            90: "% meses HR 90%",
            100: "% meses HR 100%",
        }
    )

    operational = operational.merge(dist_pivot, on="Estrategia", how="left")
    operational["Tipo"] = operational["Estrategia"].apply(_strategy_type)
    operational["Comentario"] = operational["Estrategia"].apply(_strategy_comment)

    operational = _sort_strategies(
        operational.rename(columns={"Estrategia": "strategy"}), col="strategy"
    ).rename(columns={"strategy": "Estrategia"})

    percent_cols = [c for c in operational.columns if c.startswith("% meses")]
    operational[percent_cols] = operational[percent_cols] * 100

    operational = _format_number_columns(operational, decimals=2)
    operational.to_csv(TABLES_DIR / "tabla_3_perfil_operativo_cobertura.csv", index=False)

    # Distribución simple de hedge ratio
    dist_counts_presentable = dist_counts.copy()
    dist_counts_presentable["Porcentaje_meses"] = dist_counts_presentable["Porcentaje_meses"] * 100
    dist_counts_presentable = dist_counts_presentable.rename(
        columns={
            "strategy": "Estrategia",
            "hedge_ratio_pct": "Hedge ratio (%)",
            "Meses_en_tramo": "Meses",
            "Total_meses": "Total meses",
            "Porcentaje_meses": "% de meses",
        }
    )
    dist_counts_presentable = _format_number_columns(dist_counts_presentable, decimals=2)
    dist_counts_presentable.to_csv(TABLES_DIR / "hedge_ratio_distribution.csv", index=False)

    return executive, operational, dist_counts_presentable


def _export_dl_metric_tables(lstm_metrics: pd.DataFrame, gru_metrics: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    dl_metrics = pd.concat([lstm_metrics, gru_metrics], ignore_index=True)
    dl_metrics.to_csv(TABLES_DIR / "dl_metrics_summary.csv", index=False)

    avg_metrics = (
        dl_metrics.groupby("model")
        .agg(
            MAE_promedio=("mae", "mean"),
            RMSE_promedio=("rmse", "mean"),
            MAE_minimo=("mae", "min"),
            MAE_maximo=("mae", "max"),
            RMSE_minimo=("rmse", "min"),
            RMSE_maximo=("rmse", "max"),
            Numero_splits=("split_id", "nunique"),
        )
        .reset_index()
        .rename(columns={"model": "Modelo"})
    )

    avg_metrics = _format_number_columns(avg_metrics, decimals=6)
    avg_metrics.to_csv(TABLES_DIR / "tabla_2_metricas_predictivas_promedio.csv", index=False)

    dl_metrics_presentable = dl_metrics.copy().rename(
        columns={
            "split_id": "Split",
            "model": "Modelo",
            "train_years": "Años entrenamiento",
            "val_years": "Años validación",
            "test_years": "Año test",
            "mae": "MAE",
            "rmse": "RMSE",
        }
    )
    dl_metrics_presentable = _format_number_columns(dl_metrics_presentable, decimals=6)
    dl_metrics_presentable.to_csv(TABLES_DIR / "metricas_predictivas_por_split.csv", index=False)

    return avg_metrics, dl_metrics_presentable

def _format_thousands(x, _):
    return f"{x:,.0f}"


def _apply_common_plot_format(ax, legend=True):
    ax.grid(True, alpha=0.3)

    if legend:
        ax.legend(frameon=True)

    ax.tick_params(axis="x", rotation=0)

def _plot_cumulative_pnl(backtest_df: pd.DataFrame) -> None:
    tmp = _prepare_backtest(backtest_df)

    plt.figure(figsize=(12, 6))
    ax = plt.gca()

    strategy_order = [s for s in ALL_STRATEGY_ORDER if s in tmp["strategy"].unique()]

    for strategy in strategy_order:
        group = tmp[tmp["strategy"] == strategy]
        ax.plot(group["month_dt"], group["cum_net_pnl"], label=strategy, linewidth=1.8)

    ax.set_title("PnL proxy neto acumulado por estrategia")
    ax.set_xlabel("Mes")
    ax.set_ylabel("PnL proxy neto acumulado")
    ax.yaxis.set_major_formatter(FuncFormatter(_format_thousands))

    _apply_common_plot_format(ax)

    plt.tight_layout()
    plt.savefig(FIGURES_DIR / "figura_1_pnl_acumulado_estrategias.png", dpi=220)
    plt.savefig(FIGURES_DIR / "cum_pnl_by_strategy.png", dpi=220)
    plt.close()

def _plot_cumulative_pnl_main_strategies(backtest_df: pd.DataFrame) -> None:
    tmp = _prepare_backtest(backtest_df)
    tmp = tmp[tmp["strategy"].isin(MAIN_STRATEGIES)].copy()

    plt.figure(figsize=(12, 6))

    for strategy in MAIN_STRATEGIES:
        group = tmp[tmp["strategy"] == strategy]
        if not group.empty:
            plt.plot(group["month_dt"], group["cum_net_pnl"], label=strategy)

    plt.title("PnL proxy neto acumulado por estrategia")
    plt.xlabel("Mes")
    plt.ylabel("PnL proxy neto acumulado")
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(FIGURES_DIR / "figura_1b_pnl_acumulado_estrategias_principales.png", dpi=220)
    plt.close()


def _plot_hedge_ratio_heatmap(backtest_df: pd.DataFrame) -> None:
    tmp = _prepare_backtest(backtest_df)
    tmp = tmp[tmp["strategy"].isin(DYNAMIC_STRATEGIES)].copy()

    if tmp.empty:
        return

    pivot = tmp.pivot_table(
        index="strategy",
        columns="month_dt",
        values="hedge_ratio",
        aggfunc="mean",
    )

    pivot = pivot.reindex([s for s in DYNAMIC_STRATEGIES if s in pivot.index])

    fig, ax = plt.subplots(figsize=(14, 4.5))
    im = ax.imshow(pivot.values, aspect="auto", vmin=0.3, vmax=0.9)

    ax.set_title("Mapa de calor del hedge ratio dinámico")
    ax.set_xlabel("Mes")
    ax.set_ylabel("Estrategia")

    ax.set_yticks(np.arange(len(pivot.index)))
    ax.set_yticklabels(pivot.index)

    months = list(pivot.columns)
    tick_step = max(1, len(months) // 10)
    tick_positions = np.arange(0, len(months), tick_step)

    ax.set_xticks(tick_positions)
    ax.set_xticklabels(
        [months[i].strftime("%Y-%m") for i in tick_positions],
        rotation=45,
        ha="right",
    )

    cbar = fig.colorbar(im, ax=ax)
    cbar.set_label("Hedge ratio")
    cbar.set_ticks([0.3, 0.6, 0.9])
    cbar.set_ticklabels(["30%", "60%", "90%"])

    plt.tight_layout()
    plt.savefig(FIGURES_DIR / "figura_2_heatmap_hedge_ratio_dinamico.png", dpi=220)
    plt.close()


def _plot_hedge_ratio_distribution(backtest_df: pd.DataFrame) -> None:
    tmp = _prepare_backtest(backtest_df)
    tmp = tmp[tmp["strategy"].isin(DYNAMIC_STRATEGIES)].copy()

    if tmp.empty:
        return

    dist = (
        tmp.groupby(["strategy", "hedge_ratio_pct"])
        .size()
        .reset_index(name="months")
    )

    totals = tmp.groupby("strategy")["month"].nunique().reset_index(name="total")
    dist = dist.merge(totals, on="strategy", how="left")
    dist["pct"] = dist["months"] / dist["total"] * 100

    pivot = dist.pivot_table(
        index="strategy",
        columns="hedge_ratio_pct",
        values="pct",
        fill_value=0.0,
    )

    pivot = pivot.reindex([s for s in DYNAMIC_STRATEGIES if s in pivot.index])

    for col in [30, 60, 90]:
        if col not in pivot.columns:
            pivot[col] = 0.0

    pivot = pivot[[30, 60, 90]]

    ax = pivot.plot(kind="barh", stacked=True, figsize=(10, 5))

    ax.set_title("Distribución de hedge ratios dinámicos por estrategia")
    ax.set_xlabel("% de meses")
    ax.set_ylabel("Estrategia")
    ax.legend(
        title="Hedge ratio",
        labels=["30%", "60%", "90%"],
        loc="center left",
        bbox_to_anchor=(1.01, 0.5),
        frameon=True,
    )
    ax.grid(True, axis="x", alpha=0.3)

    for container in ax.containers:
        labels = []
        for value in container.datavalues:
            labels.append(f"{value:.0f}%" if value >= 6 else "")
        ax.bar_label(container, labels=labels, label_type="center", fontsize=8)

    plt.subplots_adjust(right=0.82)
    plt.tight_layout()
    plt.savefig(FIGURES_DIR / "figura_3_distribucion_hedge_ratio_dinamico.png", dpi=220)
    plt.close()


def _plot_residual_exposure(backtest_df: pd.DataFrame) -> None:
    tmp = _prepare_backtest(backtest_df)
    tmp = tmp[tmp["strategy"].isin(MAIN_STRATEGIES)].copy()

    if tmp.empty:
        return

    pivot = tmp.pivot_table(
        index="month_dt",
        columns="strategy",
        values="abs_residual_exposure",
        aggfunc="mean",
    ).sort_index()

    rolling = pivot.rolling(6, min_periods=1).mean()

    plt.figure(figsize=(12, 6))
    ax = plt.gca()

    for strategy in MAIN_STRATEGIES:
        if strategy in rolling.columns:
            ax.plot(rolling.index, rolling[strategy], label=strategy, linewidth=1.8)

    ax.set_title("Exposición residual absoluta promedio móvil 6M")
    ax.set_xlabel("Mes")
    ax.set_ylabel("Exposición residual absoluta")
    ax.yaxis.set_major_formatter(FuncFormatter(_format_thousands))

    _apply_common_plot_format(ax)

    plt.tight_layout()
    plt.savefig(FIGURES_DIR / "figura_4_exposicion_residual_movil_6m.png", dpi=220)
    plt.close()


def _plot_dl_metrics(lstm_metrics: pd.DataFrame, gru_metrics: pd.DataFrame) -> None:
    dl_metrics = pd.concat([lstm_metrics, gru_metrics], ignore_index=True)

    plt.figure(figsize=(10, 5))
    ax = plt.gca()

    for model, group in dl_metrics.groupby("model"):
        ax.plot(group["split_id"], group["mae"], marker="o", label=f"{model} MAE", linewidth=1.8)

    ax.set_title("MAE por split walk-forward - LSTM vs GRU")
    ax.set_xlabel("Split walk-forward")
    ax.set_ylabel("MAE de volatilidad mensual")
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=1.0))

    _apply_common_plot_format(ax)

    plt.tight_layout()
    plt.savefig(FIGURES_DIR / "figura_5_mae_lstm_gru.png", dpi=220)
    plt.close()

    plt.figure(figsize=(10, 5))
    ax = plt.gca()

    for model, group in dl_metrics.groupby("model"):
        ax.plot(group["split_id"], group["rmse"], marker="o", label=f"{model} RMSE", linewidth=1.8)

    ax.set_title("RMSE por split walk-forward - LSTM vs GRU")
    ax.set_xlabel("Split walk-forward")
    ax.set_ylabel("RMSE de volatilidad mensual")
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=1.0))

    _apply_common_plot_format(ax)

    plt.tight_layout()
    plt.savefig(FIGURES_DIR / "figura_6_rmse_lstm_gru.png", dpi=220)
    plt.close()


def _plot_var_forecasts_if_available() -> pd.DataFrame | None:
    forecast_files = [
        OUTPUT_DIR / "lstm_forecasts.csv",
        OUTPUT_DIR / "gru_forecasts.csv",
        OUTPUT_DIR / "benchmark_vol_forecasts.csv",
    ]

    frames = []

    for file in forecast_files:
        if file.exists():
            df = pd.read_csv(file)

            if {"date", "model", "y_pred_var_95_1m"}.issubset(df.columns):
                frames.append(df[["date", "model", "y_pred_var_95_1m"]].copy())

    if not frames:
        return None

    forecasts = pd.concat(frames, ignore_index=True)
    forecasts["date"] = pd.to_datetime(forecasts["date"])
    forecasts["month"] = forecasts["date"].dt.to_period("M").astype(str)

    monthly = (
        forecasts.sort_values("date")
        .groupby(["month", "model"], as_index=False)
        .last()
    )

    monthly["month_dt"] = pd.PeriodIndex(monthly["month"], freq="M").to_timestamp()
    monthly = monthly[monthly["model"].isin(DYNAMIC_STRATEGIES)].copy()

    # Usar ventana común entre modelos para que la comparación visual sea consistente
    month_sets = []

    for _, group in monthly.groupby("model"):
        month_sets.append(set(group["month"].dropna().astype(str)))

    if month_sets:
        common_months = set.intersection(*month_sets)
        monthly = monthly[monthly["month"].isin(common_months)].copy()

    monthly.to_csv(TABLES_DIR / "monthly_var_forecasts_by_model.csv", index=False)

    plt.figure(figsize=(12, 6))
    ax = plt.gca()

    for model in DYNAMIC_STRATEGIES:
        group = monthly[monthly["model"] == model]
        if not group.empty:
            ax.plot(group["month_dt"], group["y_pred_var_95_1m"], label=model, linewidth=1.7)

    ax.set_title("VaR mensual proyectado por modelo")
    ax.set_xlabel("Mes")
    ax.set_ylabel("VaR mensual al 95%")
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=1.0))

    _apply_common_plot_format(ax)

    plt.tight_layout()
    plt.savefig(FIGURES_DIR / "figura_7_var_mensual_proyectado_modelo.png", dpi=220)
    plt.close()

    return monthly

def _format_excel_workbook(excel_path) -> None:
    """
    Aplica formato visual al archivo Excel de resultados:
    encabezados, autofiltro, ajuste de texto, ancho de columnas,
    congelación de paneles y formatos numéricos básicos.
    """
    wb = load_workbook(excel_path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

                if isinstance(cell.value, (int, float)):
                    header = ws.cell(row=1, column=cell.column).value
                    header_text = str(header).lower() if header is not None else ""

                    if "pnl" in header_text or "costo" in header_text or "exposición" in header_text or "exposicion" in header_text or "drawdown" in header_text:
                        cell.number_format = '#,##0.00'
                    elif "%" in header_text or "relativa" in header_text:
                        cell.number_format = '0.00'
                    elif "mae" in header_text or "rmse" in header_text or "vol" in header_text:
                        cell.number_format = '0.000000'
                    elif "hedge" in header_text:
                        cell.number_format = '0.00'
                    else:
                        cell.number_format = '#,##0.00'

        for column_cells in ws.columns:
            col_letter = get_column_letter(column_cells[0].column)
            max_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max(max_length + 2, 12), 60)
            ws.column_dimensions[col_letter].width = adjusted_width

        ws.row_dimensions[1].height = 30

    wb.save(excel_path)

def _export_excel_pack(
    executive_summary: pd.DataFrame,
    predictive_avg: pd.DataFrame,
    operational_profile: pd.DataFrame,
    hedge_distribution: pd.DataFrame,
) -> None:
    excel_path = TABLES_DIR / "resultados_tfm.xlsx"

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        executive_summary.to_excel(writer, sheet_name="1 Resumen backtest", index=False)
        predictive_avg.to_excel(writer, sheet_name="2 Metricas predictivas", index=False)
        operational_profile.to_excel(writer, sheet_name="3 Perfil cobertura", index=False)
        hedge_distribution.to_excel(writer, sheet_name="4 Distrib hedge ratio", index=False)

        vol_metrics_path = TABLES_DIR / "tabla_4_metricas_volatilidad_predicha_vs_realizada.csv"
        if vol_metrics_path.exists():
            vol_metrics = pd.read_csv(vol_metrics_path)
            vol_metrics.to_excel(writer, sheet_name="5 Vol pred vs real", index=False)

        importance_path = TABLES_DIR / "tabla_5_importancia_variables_permutation.csv"
        if importance_path.exists():
            importance = pd.read_csv(importance_path)
            importance.to_excel(writer, sheet_name="6 Importancia variables", index=False)

        traceability_path = TABLES_DIR / "tabla_6_trazabilidad_datos_variables.csv"
        if traceability_path.exists():
            traceability = pd.read_csv(traceability_path)
            traceability.to_excel(writer, sheet_name="7 Trazabilidad datos", index=False)

    _format_excel_workbook(excel_path)

def _plot_volatility_predicted_vs_realized() -> pd.DataFrame | None:
    """
    Compara la volatilidad mensual predicha por cada modelo contra la volatilidad
    realizada futura utilizada como target.
    """
    features_path = PROCESSED_DIR / "features_targets_daily.csv"

    if not features_path.exists():
        return None

    features = pd.read_csv(features_path)
    date_col = features.columns[0]

    realized = features[[date_col, "target_vol_1m"]].copy()
    realized = realized.rename(
        columns={
            date_col: "date",
            "target_vol_1m": "vol_realizada_1m",
        }
    )
    realized["date"] = pd.to_datetime(realized["date"])

    forecast_files = [
        OUTPUT_DIR / "lstm_forecasts.csv",
        OUTPUT_DIR / "gru_forecasts.csv",
        OUTPUT_DIR / "benchmark_vol_forecasts.csv",
    ]

    frames = []

    for file in forecast_files:
        if file.exists():
            df = pd.read_csv(file)

            if {"date", "model", "y_pred_vol_1m"}.issubset(df.columns):
                tmp = df[["date", "model", "y_pred_vol_1m"]].copy()
                tmp["date"] = pd.to_datetime(tmp["date"])
                frames.append(tmp)

    if not frames:
        return None

    forecasts = pd.concat(frames, ignore_index=True)
    forecasts = forecasts[forecasts["model"].isin(DYNAMIC_STRATEGIES)].copy()

    forecasts = forecasts.merge(realized, on="date", how="left")
    forecasts["month"] = forecasts["date"].dt.to_period("M").astype(str)

    monthly = (
        forecasts.sort_values("date")
        .groupby(["month", "model"], as_index=False)
        .last()
    )

    # Ventana común entre modelos
    month_sets = []

    for _, group in monthly.groupby("model"):
        month_sets.append(set(group["month"].dropna().astype(str)))

    if month_sets:
        common_months = set.intersection(*month_sets)
        monthly = monthly[monthly["month"].isin(common_months)].copy()

    monthly["month_dt"] = pd.PeriodIndex(monthly["month"], freq="M").to_timestamp()

    monthly["abs_error"] = (
        monthly["y_pred_vol_1m"] - monthly["vol_realizada_1m"]
    ).abs()

    monthly["squared_error"] = (
        monthly["y_pred_vol_1m"] - monthly["vol_realizada_1m"]
    ) ** 2

    monthly.to_csv(
        TABLES_DIR / "tabla_4_volatilidad_predicha_vs_realizada_detalle.csv",
        index=False,
    )

    metrics = (
        monthly.dropna(subset=["abs_error", "squared_error"])
        .groupby("model")
        .agg(
            MAE_volatilidad=("abs_error", "mean"),
            RMSE_volatilidad=("squared_error", lambda x: np.sqrt(x.mean())),
            Volatilidad_predicha_promedio=("y_pred_vol_1m", "mean"),
            Volatilidad_realizada_promedio=("vol_realizada_1m", "mean"),
            Meses=("month", "nunique"),
        )
        .reset_index()
        .rename(columns={"model": "Modelo"})
    )

    numeric_cols = metrics.select_dtypes(include="number").columns
    metrics[numeric_cols] = metrics[numeric_cols].round(6)

    metrics.to_csv(
        TABLES_DIR / "tabla_4_metricas_volatilidad_predicha_vs_realizada.csv",
        index=False,
    )

    realized_monthly = (
        monthly.groupby("month_dt", as_index=False)["vol_realizada_1m"]
        .first()
        .sort_values("month_dt")
    )

    plt.figure(figsize=(12, 6))
    ax = plt.gca()

    ax.plot(
        realized_monthly["month_dt"],
        realized_monthly["vol_realizada_1m"],
        label="Volatilidad realizada",
        linewidth=2.5,
        linestyle="--",
    )

    for model in DYNAMIC_STRATEGIES:
        group = monthly[monthly["model"] == model].sort_values("month_dt")

        if not group.empty:
            ax.plot(
                group["month_dt"],
                group["y_pred_vol_1m"],
                label=f"{model} predicha",
                linewidth=1.6,
            )

    ax.set_title("Volatilidad mensual predicha vs realizada")
    ax.set_xlabel("Mes")
    ax.set_ylabel("Volatilidad mensual")
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=1.0))

    _apply_common_plot_format(ax)

    plt.tight_layout()
    plt.savefig(FIGURES_DIR / "figura_8_volatilidad_predicha_vs_realizada.png", dpi=220)
    plt.close()

    return metrics

def _plot_permutation_importance_if_available() -> pd.DataFrame | None:
    """
    Genera tabla y figura de importancia por permutación para LSTM y GRU.
    La importancia se mide como el aumento del RMSE al permutar cada variable.

    Para facilitar interpretación, también se calcula una importancia relativa:
    Importancia relativa (%) = aumento RMSE / RMSE base promedio * 100
    """
    files = [
        OUTPUT_DIR / "lstm_permutation_importance.csv",
        OUTPUT_DIR / "gru_permutation_importance.csv",
    ]

    frames = []

    for file in files:
        if file.exists():
            df = pd.read_csv(file)

            required_cols = {
                "model",
                "feature",
                "importance_rmse_increase",
                "importance_rmse_std",
                "baseline_rmse",
                "n_repeats",
                "split_id",
            }

            if required_cols.issubset(df.columns):
                frames.append(df)

    if not frames:
        return None

    importance = pd.concat(frames, ignore_index=True)

    variable_labels = {
        "ret_eurusd": "Retorno EUR/USD",
        "vol_21": "Volatilidad histórica 21D",
        "ma_ratio_21": "Desviación media móvil 21D",
        "momentum_21": "Momentum EUR/USD 21D",
        "ret_vix": "Retorno VIX",
        "ret_dxy": "Retorno DXY",
        "ret_wti": "Retorno WTI",
        "US10Y": "Tasa EE. UU. 10Y",
        "EU10Y": "Tasa eurozona 10Y",
        "spread_10y": "Diferencial tasas 10Y",
    }

    summary = (
        importance.groupby(["model", "feature"], as_index=False)
        .agg(
            Aumento_RMSE_promedio=("importance_rmse_increase", "mean"),
            Desv_aumento_RMSE=("importance_rmse_increase", "std"),
            RMSE_base_promedio=("baseline_rmse", "mean"),
            Splits=("split_id", "nunique"),
        )
    )

    summary["Variable"] = summary["feature"].map(variable_labels).fillna(summary["feature"])

    # La importancia relativa se calcula sobre el aumento del RMSE.
    # Los valores negativos se conservan en el aumento absoluto, pero se truncan en 0
    # para la lectura relativa, ya que indican ausencia de contribución predictiva positiva.
    summary["Importancia_relativa_pct"] = (
        summary["Aumento_RMSE_promedio"] / summary["RMSE_base_promedio"] * 100
    )

    summary["Importancia_relativa_pct_grafica"] = summary["Importancia_relativa_pct"].clip(lower=0)

    summary = summary.sort_values(
        ["model", "Importancia_relativa_pct_grafica"],
        ascending=[True, False],
    )

    presentable = summary[
        [
            "model",
            "Variable",
            "Aumento_RMSE_promedio",
            "Desv_aumento_RMSE",
            "RMSE_base_promedio",
            "Importancia_relativa_pct",
            "Importancia_relativa_pct_grafica",
            "Splits",
        ]
    ].rename(
        columns={
            "model": "Modelo",
            "Aumento_RMSE_promedio": "Aumento promedio RMSE",
            "Desv_aumento_RMSE": "Desv. aumento RMSE",
            "RMSE_base_promedio": "RMSE base promedio",
            "Importancia_relativa_pct": "Importancia relativa (%)",
            "Importancia_relativa_pct_grafica": "Importancia relativa positiva (%)",
        }
    )

    numeric_cols = presentable.select_dtypes(include="number").columns
    presentable[numeric_cols] = presentable[numeric_cols].round(6)

    presentable.to_csv(
        TABLES_DIR / "tabla_5_importancia_variables_permutation.csv",
        index=False,
    )

    # Gráfica en importancia relativa positiva (%).
    pivot = presentable.pivot_table(
        index="Variable",
        columns="Modelo",
        values="Importancia relativa positiva (%)",
        fill_value=0.0,
    )

    pivot["Importancia_media"] = pivot.mean(axis=1)
    pivot = pivot.sort_values("Importancia_media", ascending=True)
    pivot = pivot.drop(columns="Importancia_media")

    ax = pivot.plot(kind="barh", figsize=(11, 7))

    ax.set_title("Importancia relativa de variables por permutación")
    ax.set_xlabel("Aumento relativo del RMSE (%)")
    ax.set_ylabel("Variable")
    ax.grid(True, axis="x", alpha=0.3)
    ax.legend(title="Modelo")

    plt.tight_layout()
    plt.savefig(
        FIGURES_DIR / "figura_9_importancia_variables_permutation.png",
        dpi=220,
    )
    plt.close()

    return presentable

def _export_data_traceability_table() -> pd.DataFrame:
    """
    Exporta una tabla de trazabilidad de datos, variables construidas y uso metodológico.
    La tabla resume fuente, frecuencia, tratamiento, transformación y uso en el TFM.
    """
    rows = [
        {
            "Variable": "EURUSD",
            "Fuente / ticker": "Yahoo Finance / EURUSD=X",
            "Frecuencia": "Diaria",
            "Tratamiento": "Descarga con yfinance, alineación por fecha, conversión numérica, forward fill y eliminación de nulos.",
            "Transformación": "Retorno logarítmico, media móvil 21D, momentum 21D, desviación frente a media móvil y volatilidad histórica.",
            "Uso metodológico": "Variable principal de riesgo cambiario y base para estimar volatilidad, VaR y desempeño de la cobertura.",
        },
        {
            "Variable": "VIX",
            "Fuente / ticker": "Yahoo Finance / ^VIX",
            "Frecuencia": "Diaria",
            "Tratamiento": "Alineación con el calendario de EUR/USD, forward fill y eliminación de nulos.",
            "Transformación": "Retorno logarítmico.",
            "Uso metodológico": "Proxy de aversión global al riesgo y condiciones de estrés financiero.",
        },
        {
            "Variable": "DXY",
            "Fuente / ticker": "Yahoo Finance / DX-Y.NYB",
            "Frecuencia": "Diaria",
            "Tratamiento": "Alineación con el calendario de EUR/USD, forward fill y eliminación de nulos.",
            "Transformación": "Retorno logarítmico.",
            "Uso metodológico": "Proxy de fortaleza relativa del dólar estadounidense.",
        },
        {
            "Variable": "WTI",
            "Fuente / ticker": "Yahoo Finance / CL=F",
            "Frecuencia": "Diaria",
            "Tratamiento": "Alineación con el calendario de EUR/USD, forward fill y eliminación de nulos.",
            "Transformación": "Variación porcentual.",
            "Uso metodológico": "Variable macrofinanciera asociada a commodities y condiciones globales de mercado.",
        },
        {
            "Variable": "US10Y",
            "Fuente / ticker": "FRED / DGS10",
            "Frecuencia": "Diaria",
            "Tratamiento": "Descarga mediante pandas_datareader, alineación con el resto de variables, forward fill y eliminación de nulos.",
            "Transformación": "Nivel de tasa.",
            "Uso metodológico": "Proxy de condiciones monetarias y financieras de Estados Unidos.",
        },
        {
            "Variable": "EU10Y",
            "Fuente / ticker": "ECB Data API / B.U2.EUR.4F.G_N.A.SV_C_YM.SR_10Y",
            "Frecuencia": "Diaria",
            "Tratamiento": "Descarga desde ECB Data API, conversión numérica, alineación con el resto de variables, forward fill y eliminación de nulos.",
            "Transformación": "Nivel de tasa.",
            "Uso metodológico": "Proxy de condiciones de tasas soberanas de la eurozona a 10 años.",
        },
        {
            "Variable": "spread_10y",
            "Fuente / ticker": "Variable construida",
            "Frecuencia": "Diaria",
            "Tratamiento": "Construida después de alinear US10Y y EU10Y.",
            "Transformación": "US10Y - EU10Y.",
            "Uso metodológico": "Diferencial de tasas de largo plazo entre Estados Unidos y eurozona.",
        },
        {
            "Variable": "vol_21",
            "Fuente / ticker": "Variable construida a partir de EUR/USD",
            "Frecuencia": "Diaria",
            "Tratamiento": "Ventana móvil de 21 días hábiles.",
            "Transformación": "Desviación estándar móvil de retornos del EUR/USD.",
            "Uso metodológico": "Indicador de volatilidad histórica usado como variable explicativa de LSTM/GRU.",
        },
        {
            "Variable": "ma_ratio_21",
            "Fuente / ticker": "Variable construida a partir de EUR/USD",
            "Frecuencia": "Diaria",
            "Tratamiento": "Media móvil de 21 días hábiles.",
            "Transformación": "EURUSD / media móvil 21D - 1.",
            "Uso metodológico": "Indicador técnico de desviación del tipo de cambio frente a su tendencia reciente.",
        },
        {
            "Variable": "momentum_21",
            "Fuente / ticker": "Variable construida a partir de EUR/USD",
            "Frecuencia": "Diaria",
            "Tratamiento": "Rezago de 21 días hábiles.",
            "Transformación": "EURUSD / EURUSD rezagado 21D - 1.",
            "Uso metodológico": "Indicador técnico de momentum del tipo de cambio.",
        },
        {
            "Variable": "target_vol_1m",
            "Fuente / ticker": "Variable construida",
            "Frecuencia": "Diaria con horizonte futuro de 21 días hábiles",
            "Tratamiento": "Calculada con retornos futuros del EUR/USD.",
            "Transformación": "Desviación estándar futura de 21 retornos diarios escalada por raíz de 21.",
            "Uso metodológico": "Variable objetivo para entrenamiento y evaluación de LSTM/GRU.",
        },
        {
            "Variable": "target_var_95_1m",
            "Fuente / ticker": "Variable construida",
            "Frecuencia": "Diaria / mensual para señal de cobertura",
            "Tratamiento": "Calculada a partir de la volatilidad mensual esperada.",
            "Transformación": "1.645 × volatilidad mensual esperada.",
            "Uso metodológico": "VaR paramétrico mensual al 95%, usado para clasificar el riesgo bajo, medio o alto.",
        },
        {
            "Variable": "Exposición neta mensual",
            "Fuente / ticker": "Simulación Monte Carlo",
            "Frecuencia": "Mensual",
            "Tratamiento": "Simulación de flujos diarios de entrada y salida, agregados a nivel mensual.",
            "Transformación": "Percentiles 5, 50 y 95; se usa la mediana como exposición base.",
            "Uso metodológico": "Base sobre la cual se aplica el hedge ratio y se define el nominal del forward.",
        },
        {
            "Variable": "Hedge ratio",
            "Fuente / ticker": "Variable construida a partir del VaR proyectado",
            "Frecuencia": "Mensual",
            "Tratamiento": "Clasificación según percentiles históricos 33 y 66 del VaR de cada modelo.",
            "Transformación": "Riesgo bajo = 30%; riesgo medio = 60%; riesgo alto = 90%.",
            "Uso metodológico": "Define el porcentaje de exposición mensual cubierta mediante forward.",
        },
    ]

    traceability = pd.DataFrame(rows)

    traceability.to_csv(
        TABLES_DIR / "tabla_6_trazabilidad_datos_variables.csv",
        index=False,
    )

    return traceability

def export_reports(
    backtest_df: pd.DataFrame,
    lstm_metrics: pd.DataFrame,
    gru_metrics: pd.DataFrame,
) -> None:
    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    FIGURES_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    backtest_df = backtest_df.copy()
    backtest_df["abs_residual_exposure"] = backtest_df["residual_exposure"].abs()

    # Output base mensual
    backtest_df.to_csv(OUTPUT_DIR / "backtest_results_monthly.csv", index=False)

    # Tablas principales
    executive_summary, operational_profile, hedge_distribution = _export_backtest_tables(backtest_df)
    predictive_avg, _ = _export_dl_metric_tables(lstm_metrics, gru_metrics)

    # Gráficas principales de resultados
    _plot_cumulative_pnl(backtest_df)
    _plot_hedge_ratio_heatmap(backtest_df)
    _plot_hedge_ratio_distribution(backtest_df)
    _plot_residual_exposure(backtest_df)

    # Gráficas técnicas secundarias
    _plot_dl_metrics(lstm_metrics, gru_metrics)
    _plot_var_forecasts_if_available()
    _plot_volatility_predicted_vs_realized()
    _plot_permutation_importance_if_available()

    # Tablas metodológicas adicionales
    _export_data_traceability_table()

    # Excel consolidado para presentar
    _export_excel_pack(
        executive_summary=executive_summary,
        predictive_avg=predictive_avg,
        operational_profile=operational_profile,
        hedge_distribution=hedge_distribution,
    )
